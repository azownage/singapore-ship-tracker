"""
Singapore AIS Tracker with S&P Maritime Risk Intelligence
Real-time vessel tracking with S&P Maritime compliance screening
v8.1 - All compliance indicators, standardized legend, improved UI
"""

import streamlit as st
import asyncio
import websockets
import json
from datetime import datetime, timezone, timedelta
import pandas as pd
import pydeck as pdk
from collections import defaultdict
import time
import requests
from typing import List, Dict, Optional, Tuple
import pickle
import os
import math

st.set_page_config(page_title="Singapore Ship Risk Tracker", page_icon="🚢", layout="wide")

# Constants
SGT = timezone(timedelta(hours=8))
STORAGE_FILE = "ship_data_cache.pkl"
RISK_DATA_FILE = "risk_data_cache.pkl"
MMSI_IMO_CACHE_FILE = "mmsi_imo_cache.pkl"
VESSEL_POSITION_FILE = "vessel_positions_cache.pkl"

# Standardized Legend:
# 🔴 Severe (2)
# 🟡 Warning (1)
# 🟢 Ok (0)
# ❓ Not checked / No IMO (-1)

# Navigation status codes
NAV_STATUS_NAMES = {
    0: "Under way using engine", 1: "At anchor", 2: "Not under command",
    3: "Restricted manoeuvrability", 4: "Constrained by draught", 5: "Moored",
    6: "Aground", 7: "Engaged in fishing", 8: "Under way sailing",
    9: "Reserved for HSC", 10: "Reserved for WIG", 11: "Reserved",
    12: "Reserved", 13: "Reserved", 14: "AIS-SART", 15: "Not defined"
}

def get_vessel_type_category(type_code: int) -> str:
    if type_code is None: return "Unknown"
    if 70 <= type_code <= 79: return "Cargo"
    elif 80 <= type_code <= 89: return "Tanker"
    elif 60 <= type_code <= 69: return "Passenger"
    elif type_code in [31, 32, 52]: return "Tug"
    elif type_code == 30: return "Fishing"
    elif 40 <= type_code <= 49: return "High Speed Craft"
    elif type_code == 50: return "Pilot"
    elif type_code == 51: return "SAR"
    elif type_code == 53: return "Port Tender"
    elif type_code == 55: return "Law Enforcement"
    return "Other"


def format_compliance_value(val) -> str:
    """Format compliance values with standardized emoji legend"""
    if val is None or val == -1: return "❓"  # Not checked
    elif val == 2: return "🔴"  # Severe
    elif val == 1: return "🟡"  # Warning
    elif val == 0: return "🟢"  # Ok
    return "❓"


def get_ship_color(legal_overall: int = -1) -> List[int]:
    """Get color based on compliance status"""
    colors = {
        2: [220, 53, 69, 200],    # Red - Severe
        1: [255, 193, 7, 200],    # Yellow - Warning
        0: [40, 167, 69, 200],    # Green - Ok
    }
    return colors.get(legal_overall, [128, 128, 128, 200])  # Gray - Not checked


def load_cache() -> Tuple[Dict, Dict, Dict, Dict]:
    """Load all cached data from disk"""
    caches = [{}, {}, {}, {}]
    files = [STORAGE_FILE, RISK_DATA_FILE, MMSI_IMO_CACHE_FILE, VESSEL_POSITION_FILE]
    for i, f in enumerate(files):
        if os.path.exists(f):
            try:
                with open(f, 'rb') as file:
                    caches[i] = pickle.load(file)
            except: pass
    return tuple(caches)


def save_cache(ship_cache: Dict, risk_cache: Dict, mmsi_imo_cache: Dict, vessel_positions: Dict = None):
    """Save all cached data to disk"""
    data = [(STORAGE_FILE, ship_cache), (RISK_DATA_FILE, risk_cache), 
            (MMSI_IMO_CACHE_FILE, mmsi_imo_cache)]
    if vessel_positions:
        data.append((VESSEL_POSITION_FILE, vessel_positions))
    for f, d in data:
        try:
            with open(f, 'wb') as file:
                pickle.dump(d, file)
        except: pass


def format_datetime(dt_string: str) -> str:
    """Format datetime string to readable format in SGT"""
    if not dt_string or dt_string == 'Unknown': return 'Unknown'
    try:
        if isinstance(dt_string, str):
            dt = datetime.fromisoformat(dt_string.replace('Z', '+00:00'))
        else:
            dt = dt_string
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=SGT)
        else:
            dt = dt.astimezone(SGT)
        return dt.strftime('%d %b %Y, %I:%M %p')
    except:
        return dt_string


# Initialize session state
if 'ship_static_cache' not in st.session_state:
    ship_cache, risk_cache, mmsi_imo_cache, vessel_positions = load_cache()
    st.session_state.ship_static_cache = ship_cache
    st.session_state.risk_data_cache = risk_cache
    st.session_state.mmsi_to_imo_cache = mmsi_imo_cache
    st.session_state.vessel_positions = vessel_positions
    st.session_state.last_save = time.time()

for key, default in [('selected_vessel', None), ('map_center', {"lat": 1.28, "lon": 103.85}),
                     ('show_details_imo', None), ('show_details_name', ''), ('zoom_mode', 'dot')]:
    if key not in st.session_state:
        st.session_state[key] = default


# Maritime zones loader
@st.cache_data
def load_maritime_zones(excel_path: str) -> Dict[str, List]:
    """Load maritime zones from Excel file"""
    zones = {"Anchorages": [], "Channels": [], "Fairways": []}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        for sheet_name in wb.sheetnames:
            zone_type = None
            if 'anchorage' in sheet_name.lower(): zone_type = 'Anchorages'
            elif 'channel' in sheet_name.lower(): zone_type = 'Channels'
            elif 'fairway' in sheet_name.lower(): zone_type = 'Fairways'
            if not zone_type: continue
            
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            current_zone = None
            
            for row in rows:
                if len(row) >= 4:
                    zone_name, point_num, lat, lon = row[0], row[1], row[2], row[3]
                    if zone_name and str(zone_name).strip():
                        if current_zone and current_zone['coordinates']:
                            zones[zone_type].append(current_zone)
                        current_zone = {'name': str(zone_name).strip(), 'coordinates': []}
                    if current_zone and lat and lon:
                        try:
                            current_zone['coordinates'].append([float(lon), float(lat)])
                        except: pass
            if current_zone and current_zone['coordinates']:
                zones[zone_type].append(current_zone)
    except Exception as e:
        pass
    return zones


def create_zone_layer(zones: List[Dict], color: List[int], layer_id: str) -> pdk.Layer:
    """Create a polygon layer for maritime zones"""
    polygons = []
    for zone in zones:
        if zone['coordinates'] and len(zone['coordinates']) >= 3:
            coords = zone['coordinates']
            if coords[0] != coords[-1]:
                coords = coords + [coords[0]]
            polygons.append({'polygon': coords, 'name': zone['name']})
    if not polygons: return None
    return pdk.Layer('PolygonLayer', data=polygons, get_polygon='polygon',
                     get_fill_color=color, get_line_color=[100, 100, 100, 200],
                     line_width_min_pixels=1, pickable=True)


# API response field mapping (in PDF order)
# These are the fields actually returned by the S&P API
API_COMPLIANCE_FIELDS = {
    # Field name in our code: (API field name, Table column header, Description)
    'legal_overall': ('legalOverall', 'Legal', 'Overall compliance status'),
    # Ship Sanctions (PDF Page 1)
    'ship_bes': ('shipBESSanctionList', 'UK', 'HM Treasury OFSI (UK)'),
    'ship_eu': ('shipEUSanctionList', 'EU', 'EU Sanction List'),
    'ship_ofac': ('shipOFACSanctionList', 'OFAC', 'US Treasury OFAC SDN'),
    'ship_ofac_non_sdn': ('shipOFACNonSDNSanctionList', 'OFAC-NS', 'OFAC Non-SDN'),
    'ship_swiss': ('shipSwissSanctionList', 'Swiss', 'Swiss Sanction List'),
    'ship_un': ('shipUNSanctionList', 'UN', 'UN Security Council'),
    # Page 2
    'ship_ofac_advisory': ('shipOFACAdvisoryList', 'OFACAdv', 'OFAC Advisory'),
    'port_3m': ('shipSanctionedCountryPortCallLast3m', 'Port3m', 'Sanctioned port call 3m'),
    'port_6m': ('shipSanctionedCountryPortCallLast6m', 'Port6m', 'Sanctioned port call 6m'),
    'port_12m': ('shipSanctionedCountryPortCallLast12m', 'Port12m', 'Sanctioned port call 12m'),
    'dark_activity': ('shipDarkActivityIndicator', 'Dark', 'Dark activity indicator'),
    # Page 3
    'sts_partner': ('shipSTSPartnerNonComplianceLast12m', 'STS', 'STS partner non-compliance'),
    'flag_disputed': ('shipFlagDisputed', 'FlagDisp', 'Flag disputed'),
    'flag_sanctioned': ('shipFlagSanctionedCountry', 'FlagSanc', 'Flag sanctioned country'),
    'flag_historical': ('shipHistoricalFlagSanctionedCountry', 'FlagHist', 'Historical flag sanctioned'),
    'security_event': ('shipSecurityLegalDisputeEvent', 'SecEvt', 'Security/legal dispute'),
    'not_maintained': ('shipDetailsNoLongerMaintained', 'NoMaint', 'Details no longer maintained'),
    # Owner Sanctions (Pages 3-5)
    'owner_australian': ('shipOwnerAustralianSanctionList', 'OwnAU', 'Owner Australian'),
    'owner_bes': ('shipOwnerBESSanctionList', 'OwnUK', 'Owner UK'),
    'owner_canadian': ('shipOwnerCanadianSanctionList', 'OwnCA', 'Owner Canadian'),
    'owner_eu': ('shipOwnerEUSanctionList', 'OwnEU', 'Owner EU'),
    'owner_fatf': ('shipOwnerFATFJurisdiction', 'OwnFATF', 'Owner FATF'),
    'owner_ofac_ssi': ('shipOwnerOFACSSIList', 'OwnSSI', 'Owner OFAC SSI'),
    'owner_ofac': ('shipOwnerOFACSanctionList', 'OwnOFAC', 'Owner OFAC'),
    'owner_swiss': ('shipOwnerSwissSanctionList', 'OwnSwiss', 'Owner Swiss'),
    'owner_uae': ('shipOwnerUAESanctionList', 'OwnUAE', 'Owner UAE'),
    'owner_un': ('shipOwnerUNSanctionList', 'OwnUN', 'Owner UN'),
    'owner_ofac_country': ('shipOwnerOFACSanctionedCountry', 'OwnOFACCty', 'Owner OFAC country'),
    'owner_historical_ofac': ('shipOwnerHistoricalOFACSanctionedCountry', 'OwnHistOFAC', 'Owner historical OFAC'),
    'owner_parent_nc': ('shipOwnerParentCompanyNonCompliance', 'ParentNC', 'Parent non-compliance'),
    'owner_parent_ofac': ('shipOwnerParentOFACSanctionedCountry', 'ParentOFAC', 'Parent OFAC country'),
    'owner_parent_fatf': ('shipOwnerParentFATFJurisdiction', 'ParentFATF', 'Parent FATF'),
}


class SPShipsAPI:
    """S&P Ships API for MMSI to IMO lookup"""
    def __init__(self, username: str, password: str):
        self.username = username
        self.password = password
        self.base_url = "https://shipsapi.maritime.spglobal.com/MaritimeWCF/APSShipService.svc/RESTFul"
    
    def get_imo_by_mmsi(self, mmsi: str) -> Optional[str]:
        try:
            url = f"{self.base_url}/GetShipDataByMMSI?MMSI={mmsi}"
            response = requests.get(url, auth=(self.username, self.password), timeout=15)
            if response.status_code == 200:
                data = response.json()
                ships = data.get('Ships', [])
                if ships:
                    imo = ships[0].get('LRIMOShipNo')
                    if imo and str(imo) != '0':
                        return str(imo)
        except: pass
        return None
    
    def batch_get_imo_by_mmsi(self, mmsi_list: List[str]) -> Dict[str, str]:
        results = {}
        cache = st.session_state.mmsi_to_imo_cache
        uncached = [m for m in mmsi_list if m not in cache or cache.get(m) is None]
        
        for mmsi in mmsi_list:
            if mmsi in cache and cache[mmsi]:
                results[mmsi] = cache[mmsi]
        
        if not uncached: return results
        
        progress = st.progress(0, text=f"Looking up IMO for {len(uncached)} vessels...")
        for i, mmsi in enumerate(uncached):
            imo = self.get_imo_by_mmsi(mmsi)
            if imo:
                cache[mmsi] = imo
                results[mmsi] = imo
            progress.progress((i + 1) / len(uncached))
            time.sleep(0.1)
        progress.empty()
        
        st.session_state.mmsi_to_imo_cache = cache
        save_cache(st.session_state.ship_static_cache, st.session_state.risk_data_cache, cache)
        return results


class SPMaritimeAPI:
    """S&P Maritime API for compliance screening"""
    def __init__(self, username: str, password: str):
        self.username = username
        self.password = password
        self.base_url = "https://webservices.maritime.spglobal.com/RiskAndCompliance/CompliancesByImos"
    
    def get_ship_compliance_data(self, imo_numbers: List[str]) -> Dict[str, Dict]:
        if not imo_numbers: return {}
        
        cache = st.session_state.risk_data_cache
        uncached = [imo for imo in imo_numbers if imo not in cache]
        
        if not uncached:
            return {imo: cache[imo] for imo in imo_numbers}
        
        try:
            batches = [uncached[i:i+100] for i in range(0, len(uncached), 100)]
            received = set()
            
            progress = st.progress(0, text=f"Fetching compliance for {len(uncached)} vessels...")
            for batch_idx, batch in enumerate(batches):
                url = f"{self.base_url}?imos={','.join(batch)}"
                response = requests.get(url, auth=(self.username, self.password), timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    if isinstance(data, list):
                        for ship in data:
                            imo = str(ship.get('lrimoShipNo', ''))
                            if not imo: continue
                            received.add(imo)
                            
                            # Extract all fields from API response
                            cache_entry = {'cached_at': datetime.now(SGT).isoformat()}
                            for field_name, (api_field, _, _) in API_COMPLIANCE_FIELDS.items():
                                val = ship.get(api_field)
                                if val is not None:
                                    cache_entry[field_name] = val
                            cache[imo] = cache_entry
                
                progress.progress((batch_idx + 1) / len(batches))
                time.sleep(0.3)
            progress.empty()
            
            # Mark checked but not found
            for imo in uncached:
                if imo not in received:
                    cache[imo] = {'legal_overall': 0, 'checked_but_not_found': True,
                                  'cached_at': datetime.now(SGT).isoformat()}
            
            st.session_state.risk_data_cache = cache
            save_cache(st.session_state.ship_static_cache, cache, st.session_state.get('mmsi_to_imo_cache', {}))
            
        except Exception as e:
            st.error(f"⚠️ S&P API error: {e}")
        
        return {imo: cache.get(imo, {}) for imo in imo_numbers}


class AISTracker:
    """AIS data collection and vessel tracking"""
    
    def __init__(self, use_cached_positions: bool = True):
        self.ships = defaultdict(lambda: {'latest_position': None, 'static_data': None})
        if use_cached_positions and 'vessel_positions' in st.session_state:
            for mmsi, data in st.session_state.vessel_positions.items():
                if mmsi != '_last_update':
                    self.ships[mmsi] = data
    
    def save_positions_to_cache(self):
        positions = dict(self.ships)
        positions['_last_update'] = datetime.now(SGT).isoformat()
        st.session_state.vessel_positions = positions
        st.session_state.last_data_update = positions['_last_update']
        save_cache(st.session_state.ship_static_cache, st.session_state.risk_data_cache,
                   st.session_state.get('mmsi_to_imo_cache', {}), positions)
    
    async def collect_data(self, duration: int = 30, api_key: str = "", bounding_box: List = None):
        if bounding_box is None:
            bounding_box = [[[0.5, 102.0], [2.5, 106.0]]]
        
        try:
            async with websockets.connect("wss://stream.aisstream.io/v0/stream") as ws:
                await ws.send(json.dumps({
                    "APIKey": api_key,
                    "BoundingBoxes": bounding_box,
                    "FilterMessageTypes": ["PositionReport", "ShipStaticData"]
                }))
                start_time = time.time()
                
                async for message_json in ws:
                    if time.time() - start_time > duration: break
                    
                    msg = json.loads(message_json)
                    msg_type = msg.get("MessageType")
                    
                    if msg_type == "PositionReport":
                        self._process_position(msg)
                    elif msg_type == "ShipStaticData":
                        self._process_static(msg)
                
                self.save_positions_to_cache()
        except Exception as e:
            st.error(f"AIS connection error: {e}")
    
    def _process_position(self, msg: Dict):
        meta = msg.get('MetaData', {})
        pos = msg.get('Message', {}).get('PositionReport', {})
        mmsi = pos.get('UserID')
        if not mmsi: return
        
        self.ships[mmsi]['latest_position'] = {
            'latitude': pos.get('Latitude'), 'longitude': pos.get('Longitude'),
            'sog': pos.get('Sog', 0), 'cog': pos.get('Cog', 0),
            'true_heading': pos.get('TrueHeading', 511),
            'nav_status': pos.get('NavigationalStatus', 15),
            'ship_name': meta.get('ShipName', 'Unknown'),
            'timestamp': datetime.now(SGT).isoformat()
        }
        self.ships[mmsi]['last_seen'] = datetime.now(SGT).isoformat()
    
    def _process_static(self, msg: Dict):
        static = msg.get('Message', {}).get('ShipStaticData', {})
        mmsi = static.get('UserID')
        if not mmsi: return
        
        dim = static.get('Dimension', {})
        imo = str(static.get('ImoNumber', 0))
        existing = st.session_state.ship_static_cache.get(str(mmsi), {})
        
        dim_a = dim.get('A', 0) or existing.get('dimension_a', 0) or 0
        dim_b = dim.get('B', 0) or existing.get('dimension_b', 0) or 0
        dim_c = dim.get('C', 0) or existing.get('dimension_c', 0) or 0
        dim_d = dim.get('D', 0) or existing.get('dimension_d', 0) or 0
        
        if imo == '0' and existing.get('imo', '0') != '0':
            imo = existing.get('imo')
        
        info = {
            'name': static.get('Name') or existing.get('name', 'Unknown'),
            'imo': imo, 'type': static.get('Type') or existing.get('type'),
            'dimension_a': dim_a, 'dimension_b': dim_b, 'dimension_c': dim_c, 'dimension_d': dim_d,
            'length': dim_a + dim_b, 'width': dim_c + dim_d,
            'destination': static.get('Destination') or existing.get('destination', 'Unknown'),
            'call_sign': static.get('CallSign') or existing.get('call_sign', ''),
            'cached_at': datetime.now(SGT).isoformat()
        }
        
        self.ships[mmsi]['static_data'] = info
        st.session_state.ship_static_cache[str(mmsi)] = info
    
    def get_dataframe(self, sp_api=None, ships_api=None, expiry_hours=None) -> pd.DataFrame:
        """Get vessel dataframe with compliance data"""
        data = []
        now = datetime.now(SGT)
        
        for mmsi, ship in self.ships.items():
            pos = ship.get('latest_position')
            if not pos or not pos.get('latitude'): continue
            
            # Check expiry
            if expiry_hours:
                last_seen = ship.get('last_seen')
                if last_seen:
                    try:
                        hours = (now - datetime.fromisoformat(last_seen)).total_seconds() / 3600
                        if hours > expiry_hours: continue
                    except: pass
            
            static = ship.get('static_data') or st.session_state.ship_static_cache.get(str(mmsi), {})
            
            heading = pos.get('true_heading', 511)
            if heading == 511: heading = pos.get('cog', 0)
            
            length = static.get('length', 0) or 0
            width = static.get('width', 0) or 0
            has_dims = length > 0 and width > 0
            if not has_dims: length, width = 50, 10
            
            row = {
                'mmsi': mmsi, 'name': (static.get('name') or pos.get('ship_name') or 'Unknown').strip(),
                'imo': str(static.get('imo', '0')),
                'latitude': pos.get('latitude'), 'longitude': pos.get('longitude'),
                'speed': pos.get('sog', 0), 'course': pos.get('cog', 0), 'heading': heading,
                'nav_status': pos.get('nav_status', 15),
                'nav_status_name': NAV_STATUS_NAMES.get(pos.get('nav_status', 15), 'Unknown'),
                'type': static.get('type'), 'type_name': get_vessel_type_category(static.get('type')),
                'length': length, 'width': width, 'has_dimensions': has_dims,
                'destination': (static.get('destination') or 'Unknown').strip(),
                'has_static': bool(static.get('name')),
                'last_seen': ship.get('last_seen', ''),
                'color': get_ship_color(-1)
            }
            
            # Initialize all compliance fields to -1 (not checked)
            for field_name in API_COMPLIANCE_FIELDS.keys():
                row[field_name] = -1
            
            data.append(row)
        
        df = pd.DataFrame(data)
        if len(df) == 0: return df
        
        # Get IMOs and compliance data
        valid_imos = [imo for imo in df['imo'].unique() if imo and imo != '0']
        missing_imos = df[df['imo'] == '0']['mmsi'].astype(str).unique().tolist()
        
        # MMSI -> IMO lookup
        if missing_imos and ships_api:
            mmsi_imos = ships_api.batch_get_imo_by_mmsi(missing_imos)
            for idx, row in df.iterrows():
                found = mmsi_imos.get(str(row['mmsi']))
                if found:
                    df.at[idx, 'imo'] = found
                    if found not in valid_imos: valid_imos.append(found)
        elif missing_imos:
            cache = st.session_state.get('mmsi_to_imo_cache', {})
            for idx, row in df.iterrows():
                found = cache.get(str(row['mmsi']))
                if found:
                    df.at[idx, 'imo'] = found
                    if found not in valid_imos: valid_imos.append(found)
        
        # Get compliance data
        compliance_cache = st.session_state.get('risk_data_cache', {})
        compliance_data = sp_api.get_ship_compliance_data(valid_imos) if sp_api and valid_imos else {
            imo: compliance_cache.get(imo, {}) for imo in valid_imos
        }
        
        # Apply compliance to dataframe
        for idx, row in df.iterrows():
            comp = compliance_data.get(str(row['imo']), {})
            if comp:
                for field_name in API_COMPLIANCE_FIELDS.keys():
                    val = comp.get(field_name, -1)
                    if val is not None:
                        df.at[idx, field_name] = int(val)
                
                legal = df.at[idx, 'legal_overall']
                df.at[idx, 'color'] = get_ship_color(legal)
        
        return df


def create_vessel_polygon(lat, lon, heading, length, width, scale_factor=1.5):
    """Create ship polygon with pointed bow"""
    bow_offset = length * 0.5 * scale_factor
    stern_offset = -length * 0.5 * scale_factor
    half_width = width / 2 * scale_factor
    bow_taper = length * 0.15 * scale_factor
    
    points = [
        (bow_offset, 0),
        (bow_offset - bow_taper, half_width * 0.7),
        (stern_offset + length * 0.1 * scale_factor, half_width),
        (stern_offset, half_width * 0.9),
        (stern_offset, -half_width * 0.9),
        (stern_offset + length * 0.1 * scale_factor, -half_width),
        (bow_offset - bow_taper, -half_width * 0.7),
    ]
    
    heading_rad = math.radians(heading)
    m_per_deg_lat = 111320
    m_per_deg_lon = m_per_deg_lat * math.cos(math.radians(lat))
    
    polygon = []
    for dx, dy in points:
        rot_x = dx * math.cos(heading_rad) - dy * math.sin(heading_rad)
        rot_y = dx * math.sin(heading_rad) + dy * math.cos(heading_rad)
        new_lat = lat + (rot_x / m_per_deg_lat)
        new_lon = lon + (rot_y / m_per_deg_lon)
        polygon.append([new_lon, new_lat])
    polygon.append(polygon[0])
    return polygon


def create_vessel_layers(df: pd.DataFrame, use_actual_shapes: bool = False) -> List[pdk.Layer]:
    """Create vessel map layers - dots or actual shapes"""
    if len(df) == 0: return []
    
    vessel_data = []
    for _, row in df.iterrows():
        length = row['length'] if 0 < row['length'] < 500 else 50
        width = row['width'] if 0 < row['width'] < 80 else 10
        
        legal_emoji = format_compliance_value(row['legal_overall'])
        dim_text = f"{length:.0f}m x {width:.0f}m" + ("" if row['has_dimensions'] else " (est.)")
        
        tooltip = (f"<b>{row['name']}</b><br/>IMO: {row['imo']}<br/>MMSI: {row['mmsi']}<br/>"
                   f"Type: {row['type_name']}<br/>Size: {dim_text}<br/>"
                   f"Speed: {row['speed']:.1f} kts<br/>Status: {row['nav_status_name']}<br/>"
                   f"Compliance: {legal_emoji}")
        
        vessel_data.append({
            'latitude': row['latitude'], 'longitude': row['longitude'],
            'name': row['name'], 'tooltip': tooltip, 'color': row['color'],
            'heading': row['heading'], 'length': length, 'width': width
        })
    
    if not use_actual_shapes:
        # Dot view - colored dots
        return [pdk.Layer('ScatterplotLayer', data=vessel_data,
                          get_position=['longitude', 'latitude'], get_fill_color='color',
                          get_radius=500, radius_min_pixels=4, radius_max_pixels=20, pickable=True)]
    else:
        # Actual representation - ship shapes
        polygons = []
        for v in vessel_data:
            poly = create_vessel_polygon(v['latitude'], v['longitude'], v['heading'],
                                         v['length'], v['width'], scale_factor=3)
            polygons.append({'polygon': poly, 'tooltip': v['tooltip'], 'color': v['color']})
        return [pdk.Layer('PolygonLayer', data=polygons, get_polygon='polygon',
                          get_fill_color='color', get_line_color=[50, 50, 50, 100],
                          line_width_min_pixels=1, pickable=True)]


# ==================== MAIN APP ====================

st.title("🚢 Singapore Strait Ship Risk Tracker")
st.caption("Real-time vessel tracking with S&P Maritime compliance screening")

# Control buttons and last update - right after title
col1, col2, col3, col4 = st.columns([1, 1, 1, 2])

# Sidebar setup first to get credentials
st.sidebar.title("⚙️ Settings")

# API Credentials
try:
    sp_username = st.secrets["sp_maritime"]["username"]
    sp_password = st.secrets["sp_maritime"]["password"]
    ais_api_key = st.secrets.get("aisstream", {}).get("api_key", "")
    st.sidebar.success("🔐 Credentials loaded")
except:
    with st.sidebar.expander("🔐 API Credentials"):
        sp_username = st.text_input("S&P Username", type="password")
        sp_password = st.text_input("S&P Password", type="password")
        ais_api_key = st.text_input("AISStream API Key", type="password")

# AIS Settings
st.sidebar.header("📡 AIS Settings")
duration = st.sidebar.slider("Collection time (seconds)", 10, 120, 60)
enable_compliance = st.sidebar.checkbox("Enable S&P compliance", value=True)

coverage_options = {
    "Singapore Strait Only": [[[1.15, 103.55], [1.50, 104.10]]],
    "Singapore + Approaches": [[[1.0, 103.3], [1.6, 104.3]]],
    "Malacca to SCS": [[[0.5, 102.0], [2.5, 106.0]]],
    "Extended Malacca": [[[-0.5, 100.0], [3.0, 106.0]]],
    "Full Regional": [[[-1.0, 99.0], [4.0, 108.0]]]
}
selected_coverage = st.sidebar.selectbox("Coverage area", list(coverage_options.keys()), index=2)
coverage_bbox = coverage_options[selected_coverage]

expiry_options = {"1 hour": 1, "2 hours": 2, "4 hours": 4, "8 hours": 8, "12 hours": 12, "24 hours": 24, "Never": None}
vessel_expiry_hours = expiry_options[st.sidebar.selectbox("Vessel expiry", list(expiry_options.keys()), index=2)]
st.sidebar.caption("💡 Run multiple refreshes to accumulate vessels. Old vessels auto-expire.")

# Maritime Zones
st.sidebar.header("🗺️ Maritime Zones")
show_anchorages = st.sidebar.checkbox("Anchorages", value=True)
show_channels = st.sidebar.checkbox("Channels", value=True)
show_fairways = st.sidebar.checkbox("Fairways", value=True)

# Zoom mode - just 2 options
st.sidebar.header("🔍 Map View")
zoom_mode = st.sidebar.radio("Vessel display", ["Dot", "Actual"], horizontal=True)
use_actual_shapes = (zoom_mode == "Actual")

# Load zones
maritime_zones = {"Anchorages": [], "Channels": [], "Fairways": []}
for path in ["/mnt/project/Anchorages_Channels_Fairways_Details.xlsx",
             "/mnt/user-data/uploads/Anchorages_Channels_Fairways_Details.xlsx"]:
    if os.path.exists(path):
        maritime_zones = load_maritime_zones(path)
        break

# Filters
st.sidebar.header("🔍 Filters")

# Preset options - renamed
quick_filter = st.sidebar.radio("Preset", ["All Vessels", "Dark Fleet", "Sanctioned"], horizontal=True)

if quick_filter == "Dark Fleet":
    # Legal status: severe or warning, Sanctions: dark activity, Types: tanker or cargo
    default_compliance = ["Severe (🔴)", "Warning (🟡)"]
    default_sanctions = ["Dark Activity"]
    default_types = ["Tanker", "Cargo"]
elif quick_filter == "Sanctioned":
    # Legal status: severe, Sanctions: UN or OFAC
    default_compliance = ["Severe (🔴)"]
    default_sanctions = ["UN Sanctions", "OFAC Sanctions"]
    default_types = ["All"]
else:
    default_compliance = ["All"]
    default_sanctions = ["All"]
    default_types = ["All"]

selected_compliance = st.sidebar.multiselect("Legal Status", 
    ["All", "Severe (🔴)", "Warning (🟡)", "Ok (🟢)"], default=default_compliance)
selected_sanctions = st.sidebar.multiselect("Sanctions & Dark Activity", 
    ["All", "UN Sanctions", "OFAC Sanctions", "Dark Activity"], default=default_sanctions)
vessel_types = ["All", "Cargo", "Tanker", "Passenger", "Tug", "Fishing", "High Speed Craft", "Pilot", "SAR", "Other", "Unknown"]
selected_types = st.sidebar.multiselect("Vessel Types", vessel_types, default=default_types)
selected_nav = st.sidebar.multiselect("Nav Status", ["All"] + list(NAV_STATUS_NAMES.values()), default=["All"])
show_static_only = st.sidebar.checkbox("With static data only")

# Cache stats
st.sidebar.header("💾 Cache")
vessel_count = len([k for k in st.session_state.get('vessel_positions', {}).keys() if k != '_last_update'])
st.sidebar.info(f"**Vessels:** {vessel_count} | **Compliance:** {len(st.session_state.risk_data_cache)}")

cache_col1, cache_col2 = st.sidebar.columns(2)
if cache_col1.button("🗑️ Clear"):
    st.session_state.ship_static_cache = {}
    st.session_state.risk_data_cache = {}
    st.session_state.mmsi_to_imo_cache = {}
    st.session_state.vessel_positions = {}
    save_cache({}, {}, {}, {})
    st.rerun()
if cache_col2.button("🔄 Retry IMO"):
    st.session_state.mmsi_to_imo_cache = {}
    st.rerun()

# Legend - standardized
st.sidebar.markdown("---")
st.sidebar.markdown("""
### 🎨 Legend
**Compliance Status:**
- 🔴 Severe (2)
- 🟡 Warning (1)
- 🟢 Ok (0)
- ❓ Not checked (No IMO)

**Zones:** 🔵 Anchorages | 🟡 Channels | 🟠 Fairways
""")

# Main content placeholders
status_placeholder = st.empty()
stats_placeholder = st.empty()
map_placeholder = st.empty()
table_placeholder = st.empty()


def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    if len(df) == 0: return df
    filtered = df.copy()
    
    if selected_compliance and "All" not in selected_compliance:
        levels = {"Severe (🔴)": 2, "Warning (🟡)": 1, "Ok (🟢)": 0}
        filtered = filtered[filtered['legal_overall'].isin([levels[c] for c in selected_compliance if c in levels])]
    
    if selected_sanctions and "All" not in selected_sanctions:
        mask = pd.Series([False] * len(filtered), index=filtered.index)
        if "UN Sanctions" in selected_sanctions: mask |= (filtered['ship_un'] == 2)
        if "OFAC Sanctions" in selected_sanctions: mask |= (filtered['ship_ofac'] == 2)
        if "Dark Activity" in selected_sanctions: mask |= (filtered['dark_activity'] >= 1)
        filtered = filtered[mask]
    
    if selected_types and "All" not in selected_types:
        filtered = filtered[filtered['type_name'].isin(selected_types)]
    
    if selected_nav and "All" not in selected_nav:
        filtered = filtered[filtered['nav_status_name'].isin(selected_nav)]
    
    if show_static_only:
        filtered = filtered[filtered['has_static']]
    
    return filtered


def display_data(df: pd.DataFrame, last_update: str, is_cached: bool = False):
    """Display map and table"""
    # Stats
    with stats_placeholder:
        cols = st.columns(8)
        cols[0].metric("🚢 Total", len(df))
        cols[1].metric("⚡ Moving", len(df[df['speed'] > 1]) if len(df) else 0)
        cols[2].metric("📡 Static", int(df['has_static'].sum()) if len(df) else 0)
        cols[3].metric("🔴 Severe", len(df[df['legal_overall'] == 2]) if len(df) else 0)
        cols[4].metric("🟡 Warning", len(df[df['legal_overall'] == 1]) if len(df) else 0)
        cols[5].metric("🟢 Ok", len(df[df['legal_overall'] == 0]) if len(df) else 0)
        cols[6].metric("❓ Unknown", len(df[df['legal_overall'] == -1]) if len(df) else 0)
        cols[7].metric("📐 Dims", int(df['has_dimensions'].sum()) if len(df) else 0)
    
    # Map - initial zoom level 3
    center_lat, center_lon = 1.28, 103.85
    map_zoom = 3  # Initial zoom level
    
    if st.session_state.selected_vessel and len(df):
        vessel = df[df['mmsi'] == st.session_state.selected_vessel]
        if len(vessel):
            center_lat, center_lon = vessel.iloc[0]['latitude'], vessel.iloc[0]['longitude']
            map_zoom = 12
    
    layers = []
    if show_anchorages and maritime_zones['Anchorages']:
        layer = create_zone_layer(maritime_zones['Anchorages'], [0, 255, 255, 50], "anc")
        if layer: layers.append(layer)
    if show_channels and maritime_zones['Channels']:
        layer = create_zone_layer(maritime_zones['Channels'], [255, 255, 0, 50], "ch")
        if layer: layers.append(layer)
    if show_fairways and maritime_zones['Fairways']:
        layer = create_zone_layer(maritime_zones['Fairways'], [255, 165, 0, 50], "fw")
        if layer: layers.append(layer)
    
    layers.extend(create_vessel_layers(df, use_actual_shapes))
    
    with map_placeholder:
        st.pydeck_chart(pdk.Deck(
            map_style='mapbox://styles/mapbox/dark-v10', layers=layers,
            initial_view_state=pdk.ViewState(latitude=center_lat, longitude=center_lon, zoom=map_zoom),
            tooltip={'html': '{tooltip}', 'style': {'backgroundColor': 'steelblue', 'color': 'white'}}
        ), use_container_width=True)
    
    # Table
    with table_placeholder:
        if len(df) == 0:
            st.info("No vessels to display.")
        else:
            # Sort: Not checked (-1) < Ok (0) < Warning (1) < Severe (2) for ascending
            # For descending (most severe first): Severe > Warning > Ok > Not checked
            df['_sort'] = df['legal_overall'].map({2: 3, 1: 2, 0: 1, -1: 0}).fillna(0)
            display_df = df.sort_values(['_sort', 'name'], ascending=[False, True]).drop(columns=['_sort'])
            
            # Build table columns - only include fields that have data in API response
            # Start with core columns
            table_cols = ['name', 'imo', 'mmsi', 'type_name', 'nav_status_name']
            col_names = ['Name', 'IMO', 'MMSI', 'Type', 'Nav Status']
            
            # Add compliance columns in PDF order - only if they exist and have non -1 values
            for field_name, (_, header, _) in API_COMPLIANCE_FIELDS.items():
                if field_name in display_df.columns:
                    # Check if this field has any actual data (not all -1)
                    if (display_df[field_name] != -1).any():
                        display_df[f'{field_name}_fmt'] = display_df[field_name].apply(format_compliance_value)
                        table_cols.append(f'{field_name}_fmt')
                        col_names.append(header)
            
            table_df = display_df[table_cols].copy()
            table_df.columns = col_names
            
            selected = st.dataframe(table_df, use_container_width=True, height=400, hide_index=True,
                                    on_select="rerun", selection_mode="single-row")
            
            if selected and selected.selection and selected.selection.rows:
                idx = selected.selection.rows[0]
                mmsi = display_df.iloc[idx]['mmsi']
                st.info(f"Selected: **{table_df.iloc[idx]['Name']}** (IMO: {table_df.iloc[idx]['IMO']})")
                if st.button("🗺️ View on Map"):
                    st.session_state.selected_vessel = mmsi
                    st.rerun()


# Action buttons (in the columns we defined earlier)
with col1:
    refresh_clicked = st.button("🔄 Refresh Now", type="primary", use_container_width=True)

with col2:
    view_cached_clicked = st.button("📋 View Cached", use_container_width=True)

with col3:
    if st.button("🔄 Reset View", use_container_width=True):
        st.session_state.selected_vessel = None
        st.rerun()

with col4:
    last_update_time = st.session_state.get('last_data_update', 'Never')
    if last_update_time and last_update_time != 'Never':
        st.caption(f"📅 Last Updated: {format_datetime(last_update_time)} SGT")
    else:
        st.caption("📅 Last Updated: Never")


# Handle button clicks
if refresh_clicked:
    sp_api = SPMaritimeAPI(sp_username, sp_password) if enable_compliance and sp_username else None
    ships_api = SPShipsAPI(sp_username, sp_password) if enable_compliance and sp_username else None
    
    with status_placeholder:
        with st.spinner(f'Collecting AIS data for {duration}s...'):
            tracker = AISTracker(use_cached_positions=True)
            if ais_api_key:
                asyncio.run(tracker.collect_data(duration, ais_api_key, coverage_bbox))
            else:
                st.warning("No AISStream API key")
            df = tracker.get_dataframe(sp_api, ships_api, vessel_expiry_hours)
    
    status_placeholder.empty()
    if not df.empty:
        df = apply_filters(df)
        display_data(df, st.session_state.get('last_data_update', ''), is_cached=False)
    else:
        st.warning("No ships detected.")

elif view_cached_clicked:
    if 'vessel_positions' in st.session_state and st.session_state.vessel_positions:
        tracker = AISTracker(use_cached_positions=True)
        df = tracker.get_dataframe(sp_api=None, ships_api=None, expiry_hours=vessel_expiry_hours)
        df = apply_filters(df)
        display_data(df, st.session_state.vessel_positions.get('_last_update', ''), is_cached=True)
    else:
        st.info("No cached data. Click 'Refresh Now' first.")
else:
    # Auto-display cached on load
    if 'vessel_positions' in st.session_state and st.session_state.vessel_positions:
        tracker = AISTracker(use_cached_positions=True)
        df = tracker.get_dataframe(sp_api=None, ships_api=None, expiry_hours=vessel_expiry_hours)
        df = apply_filters(df)
        if not df.empty:
            display_data(df, st.session_state.vessel_positions.get('_last_update', ''), is_cached=True)
        else:
            st.info("No vessels match filters. Adjust filters or click 'Refresh Now'.")
    else:
        st.info("👆 Click 'Refresh Now' to start collecting AIS data")
